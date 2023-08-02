import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
def automate_chrome(url):
    chrome_options = Options()
    # Uncomment the next line to run Chrome in headless mode (no visible window)
    # chrome_options.add_argument("--headless")

    driver = webdriver.Chrome(options=chrome_options)
    driver.get(url)  # Change this URL if needed

    # time.sleep(5)

    # Find the button by class name
    button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Sign in')]")))

    button.click()
    # Find the input fields by ID
    email_field = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "session_key")))
    password_field = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "session_password")))
    # email_field = driver.find_element_by_id("session_key")
    # password_field = driver.find_element_by_id("session_password")
    
    # Replace "arpit.jha@ucertify" and "arp12345" with the desired email and password
    email_address = "arpit.jha@ucertify"
    password = "arp12345"
    
    email_field.send_keys(email_address)
    password_field.send_keys(password)

    button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Sign in']")))

    button.click()
    print("sign in button clicked")
    # search_box = driver.find_element_by_name("q")
    # search_box.send_keys(url)
    # search_box.submit()

    # Wait for the page to load
    time.sleep(5)

    # Retrieve the current URL from the address bar
    current_url = driver.current_url
    print("Current URL for search term '{}': {}".format(url, current_url))

    driver.quit()

def main():
    # Replace 'urls.xlsx' with the path to your Excel file containing URLs
    excel_file = 'data.xlsx'

    # Load the Excel file
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active


    # Create a new workbook to store the retrieved URLs
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active
    new_sheet.append(["Original URL", "Retrieved URL"])

    # Iterate through each row and read the URLs from the first column (Column A)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        url = row[0]
        print(row[0])
        retrieved_url = automate_chrome(url)
        new_sheet.append([url, retrieved_url])

    # Save the new workbook with retrieved URLs
    new_excel_file = 'retrieved_urls.xlsx'
    new_workbook.save(new_excel_file)
    print("Retrieved URLs saved to '{}'.".format(new_excel_file))

if __name__ == "__main__":
    main()
